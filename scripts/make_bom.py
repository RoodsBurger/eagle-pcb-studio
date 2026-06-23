#!/usr/bin/env python3
"""Generate a PCBWay-format assembly BOM .xlsx from an authoritative CSV.

Maps a flexible input CSV (tolerant, case-insensitive header detection) onto the
official PCBWay assembly BOM template. Non-populated parts are typed "DNS" and
their rows are highlighted.

Usage:
    python3 make_bom.py <input.csv> -o <output.xlsx>
"""

import argparse
import csv
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "error: openpyxl is required for this script "
        "(pip install openpyxl)\n"
    )
    sys.exit(1)


# PCBWay official template header; * marks required fields.
TEMPLATE_COLUMNS = [
    "Item #",
    "*Designator",
    "*Qty",
    "Manufacturer",
    "*Mfg Part #",
    "Description / Value",
    "*Package/Footprint",
    "Type",
    "LCSC#",
    "Your Instructions / Notes",
]

# Output columns that must carry a value for a populated row to be complete.
REQUIRED_OUTPUT_FIELDS = [
    "*Designator",
    "*Qty",
    "*Mfg Part #",
    "*Package/Footprint",
]

# Sensible per-column widths keyed by template header.
COLUMN_WIDTHS = {
    "Item #": 7,
    "*Designator": 26,
    "*Qty": 6,
    "Manufacturer": 20,
    "*Mfg Part #": 24,
    "Description / Value": 30,
    "*Package/Footprint": 18,
    "Type": 7,
    "LCSC#": 11,
    "Your Instructions / Notes": 50,
}


def normalize(name):
    """Lowercase a header and strip everything but alphanumerics for matching."""
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


# Input-header aliases (normalized) mapped to a canonical logical field.
HEADER_ALIASES = {
    "refdes": "designator",
    "designator": "designator",
    "designators": "designator",
    "reference": "designator",
    "references": "designator",
    "ref": "designator",
    "qty": "qty",
    "quantity": "qty",
    "value": "value",
    "val": "value",
    "description": "description",
    "desc": "description",
    "comment": "description",
    "package": "package",
    "footprint": "package",
    "packagefootprint": "package",
    "pattern": "package",
    "mount": "type",
    "mounttype": "type",
    "type": "type",
    "mountingtype": "type",
    "mpn": "mpn",
    "mfgpart": "mpn",
    "mfgpartnumber": "mpn",
    "manufacturerpartnumber": "mpn",
    "manufacturerpart": "mpn",
    "partnumber": "mpn",
    "mfgpartno": "mpn",
    "manufacturer": "manufacturer",
    "mfg": "manufacturer",
    "mfr": "manufacturer",
    "manufacture": "manufacturer",
    "lcsc": "lcsc",
    "lcscpartnumber": "lcsc",
    "lcscpart": "lcsc",
    "lcscno": "lcsc",
    "populate": "populate",
    "populated": "populate",
    "dnp": "dnp",
    "dns": "dnp",
    "fitted": "populate",
    "fit": "populate",
    "notes": "notes",
    "note": "notes",
    "instructions": "notes",
    "remarks": "notes",
    "comments": "notes",
}


def read_csv_rows(path):
    """Read the CSV, skipping leading comment lines, returning (headers, rows).

    Comment lines start with '#' and appear before the real header. The first
    non-comment, non-blank line is treated as the header row.
    """
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        lines = fh.readlines()

    data_lines = []
    header_seen = False
    for line in lines:
        stripped = line.strip()
        if not header_seen:
            if not stripped or stripped.startswith("#"):
                continue
            header_seen = True
        data_lines.append(line)

    if not data_lines:
        return [], []

    reader = csv.reader(data_lines)
    rows = list(reader)
    if not rows:
        return [], []
    headers = rows[0]
    body = [r for r in rows[1:] if any((c or "").strip() for c in r)]
    return headers, body


def build_field_map(headers):
    """Map each logical field to its column index in the input CSV."""
    field_to_index = {}
    for idx, header in enumerate(headers):
        canonical = HEADER_ALIASES.get(normalize(header))
        if canonical and canonical not in field_to_index:
            field_to_index[canonical] = idx
    return field_to_index


def cell(row, field_map, field):
    """Fetch a trimmed cell value for a logical field, or '' if absent."""
    idx = field_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def split_designators(raw):
    """Split a designator cell on ';' or ',' into individual refdes tokens."""
    parts = re.split(r"[;,]", raw)
    return [p.strip() for p in parts if p.strip()]


def is_populated(row, field_map):
    """Decide whether a part is populated from populate/dnp columns.

    Defaults to populated when no relevant column is present or the value is
    ambiguous.
    """
    populate = cell(row, field_map, "populate")
    if populate:
        low = populate.lower()
        if low in ("no", "n", "false", "0", "dnp", "dns", "donotplace"):
            return False
        if low in ("yes", "y", "true", "1", "place", "fitted", "fit"):
            return True
    dnp = cell(row, field_map, "dnp")
    if dnp:
        low = dnp.lower()
        if low in ("yes", "y", "true", "1", "dnp", "dns"):
            return False
        if low in ("no", "n", "false", "0"):
            return True
    return True


def normalize_type(raw, populated):
    """Map a raw mount value to SMD/THT; non-populated parts become DNS."""
    if not populated:
        return "DNS"
    low = normalize(raw)
    if low in ("smd", "smt", "surfacemount"):
        return "SMD"
    if low in ("tht", "th", "throughhole", "pth"):
        return "THT"
    if raw:
        return raw.strip().upper()
    return ""


def build_description(row, field_map):
    """Combine Description and Value into the single template field."""
    desc = cell(row, field_map, "description")
    value = cell(row, field_map, "value")
    if desc and value:
        if normalize(value) and normalize(value) in normalize(desc):
            return desc
        return f"{value} - {desc}"
    return desc or value


def map_row(row, field_map):
    """Turn one input row into a dict keyed by template column names."""
    populated = is_populated(row, field_map)
    designators = split_designators(cell(row, field_map, "designator"))

    qty_raw = cell(row, field_map, "qty")
    if qty_raw and qty_raw.isdigit():
        qty = int(qty_raw)
    else:
        qty = len(designators) if designators else (int(qty_raw) if qty_raw.isdigit() else "")

    out = {
        "*Designator": "; ".join(designators) if designators else cell(row, field_map, "designator"),
        "*Qty": qty,
        "Manufacturer": cell(row, field_map, "manufacturer"),
        "*Mfg Part #": cell(row, field_map, "mpn"),
        "Description / Value": build_description(row, field_map),
        "*Package/Footprint": cell(row, field_map, "package"),
        "Type": normalize_type(cell(row, field_map, "type"), populated),
        "LCSC#": cell(row, field_map, "lcsc"),
        "Your Instructions / Notes": cell(row, field_map, "notes"),
    }
    return out, populated, designators


def missing_required(mapped, populated):
    """Return required template fields that are empty for a populated row."""
    if not populated:
        return []
    missing = []
    for field in REQUIRED_OUTPUT_FIELDS:
        val = mapped.get(field, "")
        if val == "" or (isinstance(val, str) and not val.strip()) or val in ("-",):
            missing.append(field)
    return missing


def write_workbook(mapped_rows, out_path):
    """Write the styled PCBWay BOM workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    dns_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Header row.
    ws.append(TEMPLATE_COLUMNS)
    for col_idx, name in enumerate(TEMPLATE_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    centered_cols = {"Item #", "*Qty", "Type", "LCSC#"}

    for item_no, (mapped, populated, _desigs) in enumerate(mapped_rows, start=1):
        values = [item_no]
        for col in TEMPLATE_COLUMNS[1:]:
            values.append(mapped.get(col, ""))
        ws.append(values)
        excel_row = item_no + 1
        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            c = ws.cell(row=excel_row, column=col_idx)
            c.border = border
            c.alignment = center if col_name in centered_cols else wrap
            if not populated:
                c.fill = dns_fill

    # Column widths.
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 14)

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(
        description="Generate a PCBWay assembly BOM .xlsx from a CSV."
    )
    parser.add_argument("input", help="Input BOM CSV path")
    parser.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    headers, body = read_csv_rows(args.input)
    if not headers:
        sys.stderr.write("error: no header/data rows found in input CSV\n")
        sys.exit(1)

    field_map = build_field_map(headers)
    if "designator" not in field_map:
        sys.stderr.write(
            "error: could not detect a Designator/RefDes column in input\n"
        )
        sys.exit(1)

    mapped_rows = []
    total_designators = 0
    dns_count = 0
    missing_lcsc = 0
    rows_missing_required = []

    for row in body:
        mapped, populated, designators = map_row(row, field_map)
        mapped_rows.append((mapped, populated, designators))
        total_designators += len(designators)
        if not populated:
            dns_count += 1
        if not mapped.get("LCSC#") or mapped.get("LCSC#") in ("-",):
            missing_lcsc += 1
        miss = missing_required(mapped, populated)
        if miss:
            rows_missing_required.append((mapped.get("*Designator", "?"), miss))

    write_workbook(mapped_rows, args.output)

    print(f"Wrote PCBWay assembly BOM: {args.output}")
    print(f"  Line items:        {len(mapped_rows)}")
    print(f"  Total designators: {total_designators}")
    print(f"  DNS (not placed):  {dns_count}")
    print(f"  Rows missing LCSC#: {missing_lcsc}")
    if rows_missing_required:
        print(f"  Rows missing a required field: {len(rows_missing_required)}")
        for desig, miss in rows_missing_required:
            print(f"    - {desig}: missing {', '.join(miss)}")
    else:
        print("  Rows missing a required field: 0")


if __name__ == "__main__":
    main()
